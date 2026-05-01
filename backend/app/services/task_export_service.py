# --- task_export_service.py ---

import io
import pandas as pd
import yaml
from sqlalchemy import and_, or_, case
from app.models import (
    db,
    Task,
    Objective,
    ProgressUpdate,
    User,
    TaskAccessUser,
    TaskAccessOrganization,
    UserTaskOrder,
)
from app.constants import TaskAccessLevelEnum, StatusEnum, STATUS_LABELS


class ProgressFormatter:
    def __init__(self, db_session):
        self.db = db_session

    def list_for_objective(self, objective_id):
        progresses = (
            self.db.session.query(ProgressUpdate)
            .filter_by(objective_id=objective_id, is_deleted=False)
            .order_by(ProgressUpdate.report_date.asc())
            .all()
        )
        result = []
        for p in progresses:
            result.append({
                "内容": p.detail or "",
                "日付": p.report_date.strftime("%Y-%m-%d") if p.report_date else "",
                "報告者": self._get_user_name(p.updated_by)
            })
        return result

    def _get_user_name(self, user_id):
        user = self.db.session.get(User, user_id)
        return user.name if user else ""


from app.constants import StatusEnum, STATUS_LABELS
from app.models import Objective, User

class ObjectiveFormatter:
    def __init__(self, db_session):
        self.db = db_session
        self.progress_formatter = ProgressFormatter(db_session)

    def list_for_task(self, task_id):
        objectives = (
            self.db.session.query(Objective)
            .filter_by(task_id=task_id, is_deleted=False)
            .order_by(Objective.display_order.asc())
            .all()
        )
        result = []
        for obj in objectives:
            result.append({
                "オブジェクティブ名": obj.title,
                "期限": obj.due_date.strftime("%Y-%m-%d") if obj.due_date else "",
                "ステータス": self._get_status_label(obj.status),
                "担当者": self._get_user_name(obj.assigned_user_id),
                "progresses": self.progress_formatter.list_for_objective(obj.id)
            })
        return result

    def _get_status_label(self, status):
        try:
            return STATUS_LABELS.get(StatusEnum(status), "")
        except ValueError:
            return ""

    def _get_user_name(self, user_id):
        if not user_id:
            return ""  # ユーザー未割り当てなら空文字を返す
        user = self.db.session.get(User, user_id)
        return user.name if user else ""



class TaskDataExporter:
    def __init__(self, user_id, db_session):
        self.user_id = user_id
        self.db = db_session
        self.objective_formatter = ObjectiveFormatter(db_session)

    def get_user(self):
        return self.db.session.get(User, self.user_id)
    
    def _get_status_name(self, status):
        try:
            return STATUS_LABELS.get(StatusEnum(status), "")
        except ValueError:
            return ""

    def _get_user_name(self, user_id):
        user = self.db.session.get(User, user_id)
        return user.name if user else ""

    def get_tasks(self):
        user = self.get_user()
        if not user:
            return []

        filter_conditions = [
            Task.created_by == self.user_id,
            Task.id.in_(
                db.session.query(TaskAccessUser.task_id)
                .filter(TaskAccessUser.user_id == self.user_id)
                .filter(
                    TaskAccessUser.access_level.in_([
                        TaskAccessLevelEnum.VIEW,
                        TaskAccessLevelEnum.EDIT,
                        TaskAccessLevelEnum.FULL,
                    ])
                )
            ),
        ]

        if user.organization_id:
            filter_conditions.append(
                Task.id.in_(
                    db.session.query(TaskAccessOrganization.task_id)
                    .filter(TaskAccessOrganization.organization_id == user.organization_id)
                    .filter(
                        TaskAccessOrganization.access_level.in_([
                            TaskAccessLevelEnum.VIEW,
                            TaskAccessLevelEnum.EDIT,
                            TaskAccessLevelEnum.FULL,
                        ])
                    )
                )
            )
            # 組織に直接属するタスクも追加
            filter_conditions.append(Task.organization_id == user.organization_id)

        result = (
            db.session.query(Task, UserTaskOrder.display_order.label('user_order'))
            .outerjoin(UserTaskOrder, and_(
                UserTaskOrder.task_id == Task.id,
                UserTaskOrder.user_id == self.user_id 
            ))
            .filter(
                and_(
                    Task.is_deleted != True,
                    or_(*filter_conditions)
                )
            )
            .order_by(
                case((UserTaskOrder.display_order.is_(None), 1), else_=0),  # NULLを後ろへ
                UserTaskOrder.display_order.asc(),
                case((Task.display_order.is_(None), 1), else_=0),           # こちらもNULL後ろ
                Task.display_order.asc(),
            )
            .all()
        )

        return [tup[0] for tup in result]

    def build_nested_export_data(self):
        tasks = self.get_tasks()
        data = []
        for task in tasks:
            task_entry = {
                "タスク名": task.title,
                "期限": task.due_date.strftime("%Y-%m-%d") if task.due_date else "",
                "ステータス": self._get_status_name(task.status),
                "作成者": self._get_user_name(task.created_by),
                "objectives": self.objective_formatter.list_for_task(task.id)
            }
            data.append(task_entry)
        return data

    def export_as_excel(self):
        rows = self.build_flat_rows_for_excel()
        df = pd.DataFrame.from_records(rows)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, header=False, sheet_name='Tasks')
            self.apply_excel_formatting(writer.sheets['Tasks'])
        output.seek(0)
        return output

    def export_as_yaml(self):
        data = self.build_nested_export_data()
        return yaml.dump(data, allow_unicode=True, sort_keys=False)

    def build_flat_rows_for_excel(self):
        user = self.get_user()
        tasks = self.get_tasks()
        rows = []

        rows.append(["user", f"ユーザー名：{user.name}（ID：{user.id}）"])
        rows.append(["spacer"])

        for task in tasks:
            task_cell = f"{task.title}（期限：{task.due_date.strftime('%Y-%m-%d') if task.due_date else '未設定'}　作成者：{self._get_user_name(task.created_by)})"
            task_status = self._get_status_name(task.status)
            rows.append(["task", task_cell, task_status])
            rows.append(["header", "オブジェクティブ名", "期限", "ステータス", "担当者", "進捗内容", "進捗日", "報告者"])

            for obj in self.objective_formatter.list_for_task(task.id):
                progresses = obj["progresses"] or [{}]
                for i, p in enumerate(progresses):
                    type_prefix = "objective.multi" if i > 0 else "objective"
                    if i > 0 and len(progresses) > 2:
                        type_prefix = "objective.dotted"
                    rows.append([
                        type_prefix,
                        obj["オブジェクティブ名"] if i == 0 else "",
                        obj["期限"] if i == 0 else "",
                        obj["ステータス"] if i == 0 else "",
                        obj["担当者"] if i == 0 else "",
                        p.get("内容", ""),
                        p.get("日付", ""),
                        p.get("報告者", "")
                    ])
            rows.append(["spacer"])

        return rows

    def apply_excel_formatting(self, worksheet):
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        col_widths = [12, 68, 11, 11, 11, 57, 11, 11]
        for i, width in enumerate(col_widths, start=1):
            col_letter = worksheet.cell(row=1, column=i).column_letter
            worksheet.column_dimensions[col_letter].width = width
        worksheet.column_dimensions['A'].hidden = True
        wrap_columns = [2, 6]
        header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for row_idx in range(1, worksheet.max_row + 1):
            type_cell = worksheet.cell(row=row_idx, column=1)
            row_type = type_cell.value or ""
            for col_idx, cell in enumerate(worksheet[row_idx], start=1):
                if col_idx == 1:
                    continue
                if col_idx in wrap_columns:
                    cell.alignment = Alignment(wrap_text=True)
                if row_type == "user" and col_idx == 2:
                    cell.font = Font(size=14)
                elif row_type == "task" and col_idx in (2, 3):
                    cell.font = Font(size=14, bold=True)
                elif row_type == "header":
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                elif row_type.startswith("objective"):
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='top', wrap_text=(col_idx in wrap_columns))
